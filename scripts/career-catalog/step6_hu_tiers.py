"""Magyar relevancia-tierek + Excel review-export.

Tier-jelentés:
  T1 — nagy létszámú, gyakori magyar foglalkozás
  T2 — létezik Magyarországon, közepes/kisebb létszám (DEFAULT)
  T3 — niche: létezik, de nagyon kicsi vagy nagyon szűk szegmens
  T4 — magyar munkaerőpiacon nem értelmezhető / elhanyagolható

A tierezés szakértői becslés (Claude, 2026-07-30), a KSH nem közöl FEOR-4
szintű létszámot. A user Excelben validálja; a T1+T2 megy a termékbe.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPO = "/Users/leinadoknilik/trita/codebase"
DATA = os.path.join(REPO, "docs/product/data/occupations-v2.json")

T1 = """
11-1011.00 11-1021.00 11-2021.00 11-2022.00 11-3031.00 11-3061.00 11-3071.00 11-3121.00
11-9021.00 11-9051.00 11-9111.00 13-1031.00 13-1051.00 13-1071.00 13-1082.00 13-1111.00
13-1121.00 13-1151.00 13-1161.00 13-2011.00 13-2052.00 13-2072.00
15-1211.00 15-1212.00 15-1232.00 15-1242.00 15-1244.00 15-1251.00 15-1252.00 15-1253.00
15-1254.00 15-1299.09 15-2051.00
19-3033.00 21-1021.00 23-1011.00 25-2011.00 25-2012.00 25-2021.00 25-2022.00 25-2031.00
25-2032.00 25-3041.00 25-9042.00
27-1024.00 27-3023.00 27-3091.00
29-1021.00 29-1051.00 29-1123.00 29-1141.00 29-1215.00 29-1216.00 29-1221.00 29-1223.00
29-2034.00 29-2043.00 29-2052.00 29-2061.00 31-1121.00 31-1131.00 31-9011.00
33-2011.00 33-3051.00 33-9032.00
35-1011.00 35-2014.00 35-3011.00 35-3023.00 35-3031.00 37-2011.00 37-2012.00
39-5012.00 39-5092.00 39-5094.00 39-9011.00 39-9031.00
41-1011.00 41-2011.00 41-2031.00 41-3021.00 41-4011.00 41-4012.00 41-9022.00
43-3031.00 43-3051.00 43-4051.00 43-5011.00 43-5021.00 43-5061.00 43-5071.00 43-6011.00
43-6014.00 43-9061.00
47-2021.00 47-2031.00 47-2044.00 47-2061.00 47-2073.00 47-2081.00 47-2111.00 47-2141.00
47-2152.00 47-2181.00
49-1011.00 49-3023.00 49-3031.00 49-9021.00 49-9041.00 49-9071.00
51-1011.00 51-2022.00 51-2092.00 51-3011.00 51-3021.00 51-4041.00 51-4121.00 51-6031.00
51-9061.00 51-9111.00 51-9161.00 51-9162.00
53-3032.00 53-3033.00 53-3052.00 53-3054.00 53-7051.00 53-7062.00 53-7064.00 53-7065.00
53-7081.00
""".split()

T3 = """
11-9013.00 11-9031.00 11-9033.00 11-9071.00 11-9072.00 11-9131.00 11-9161.00 11-9171.00
11-9179.02 11-9199.08 13-1011.00 13-1021.00 13-1032.00 13-1041.04 13-1074.00 13-1199.04
13-2022.00 13-2061.00 13-2099.01 13-2099.04
15-1221.00 15-1243.01 15-1255.01 15-1299.02 15-1299.04 15-1299.06 15-1299.07 15-2011.00
15-2021.00 15-2041.01 15-2051.02 15-2099.01
17-1012.00 17-1021.00 17-1022.01 17-2011.00 17-2021.00 17-2031.00 17-2072.01 17-2111.02
17-2141.01 17-2151.00 17-2161.00 17-2171.00 17-2199.06 17-2199.07 17-2199.09
17-3021.00 17-3026.01 17-3029.08
19-1011.00 19-1013.00 19-1023.00 19-1029.01 19-1029.03 19-1031.00 19-1031.03 19-1032.00
19-1041.00 19-2011.00 19-2012.00 19-2021.00 19-2032.00 19-2042.00 19-2043.00 19-2099.01
19-3022.00 19-3032.00 19-3039.02 19-3039.03 19-3041.00 19-3091.00 19-3092.00 19-3093.00
19-3094.00 19-3099.01 19-4012.01 19-4043.00 19-4044.00 19-4051.00 19-4051.02 19-4061.00
19-4071.00 19-4092.00 19-4099.03
21-1011.00 21-1015.00 21-1092.00 21-2011.00 21-2021.00
23-1012.00 23-1021.00 23-1022.00 23-1023.00 23-2093.00
25-1011.00 25-1021.00 25-1022.00 25-1031.00 25-1032.00 25-1041.00 25-1042.00 25-1043.00
25-1051.00 25-1052.00 25-1053.00 25-1054.00 25-1061.00 25-1062.00 25-1063.00 25-1064.00
25-1065.00 25-1066.00 25-1067.00 25-1071.00 25-1072.00 25-1081.00 25-1082.00 25-1111.00
25-1112.00 25-1113.00 25-1121.00 25-1122.00 25-1123.00 25-1124.00 25-1125.00 25-1126.00
25-1192.00 25-1193.00 25-1194.00 25-2051.00 25-2055.00 25-2059.01 25-4011.00 25-4012.00
25-4013.00 25-4031.00 25-9044.00
27-1012.00 27-1013.00 27-1014.00 27-1023.00 27-1027.00 27-2011.00 27-2012.03 27-2012.04
27-2012.05 27-2021.00 27-2023.00 27-2031.00 27-2032.00 27-2041.00 27-2091.00 27-3011.00
27-3043.05 27-3092.00 27-4012.00 27-4015.00
29-1011.00 29-1022.00 29-1023.00 29-1024.00 29-1081.00 29-1122.01 29-1124.00 29-1128.00
29-1129.01 29-1129.02 29-1151.00 29-1181.00 29-1229.01 29-1229.02 29-1229.05 29-1229.06
29-1243.00 29-1291.00 29-1299.01 29-1299.02 29-2011.01 29-2011.02 29-2011.04 29-2012.01
29-2033.00 29-2036.00 29-2051.00 29-2057.00 29-2091.00 29-2092.00 29-2099.01 29-2099.05
29-2099.08 29-9091.00 29-9092.00
31-1132.00 31-1133.00 31-2011.00 31-2012.00 31-2021.00 31-2022.00 31-9093.00 31-9095.00
31-9096.00 31-9099.01 31-9099.02
33-1011.00 33-1021.00 33-2022.00 33-3011.00 33-3012.00 33-3021.02 33-3031.00 33-3041.00
33-3051.04 33-9011.00 33-9021.00 33-9091.00 33-9092.00 33-9093.00 33-9099.02
35-2012.00 35-2015.00 35-3041.00 37-3012.00 37-3013.00
39-1013.00 39-2011.00 39-3011.00 39-3092.00 39-3093.00 39-4011.00 39-4012.00 39-4021.00
39-4031.00 39-5091.00 39-6011.00 39-6012.00 39-9041.00
41-9011.00 41-9012.00 41-9091.00
43-3041.00 43-4021.00 43-4061.00 43-4121.00 43-4141.00 43-5041.00 43-9031.00 43-9071.00
43-9081.00 43-9111.00
45-2011.00 45-2021.00 45-2041.00 45-2093.00 45-3031.00 45-4011.00 45-4021.00 45-4022.00
45-4023.00
47-2011.00 47-2041.00 47-2042.00 47-2043.00 47-2053.00 47-2072.00 47-2131.00 47-2132.00
47-2142.00 47-2151.00 47-2161.00 47-2171.00 47-2221.00 47-3011.00 47-3012.00 47-3013.00
47-3014.00 47-3015.00 47-3016.00 47-4031.00 47-4061.00 47-4071.00 47-5022.00 47-5023.00
47-5032.00 47-5051.00
49-2021.00 49-2091.00 49-2093.00 49-2095.00 49-2096.00 49-2097.00 49-3022.00 49-3041.00
49-3043.00 49-3051.00 49-3052.00 49-3053.00 49-3091.00 49-3092.00 49-9011.00 49-9045.00
49-9061.00 49-9063.00 49-9064.00 49-9081.00 49-9091.00 49-9095.00 49-9096.00 49-9097.00
51-2011.00 51-2021.00 51-2051.00 51-4051.00 51-4052.00 51-4061.00 51-4062.00 51-4071.00
51-4192.00 51-4194.00 51-5111.00 51-5113.00 51-6021.00 51-6041.00 51-6042.00 51-6051.00
51-6061.00 51-6064.00 51-6091.00 51-6092.00 51-7021.00 51-7031.00 51-7032.00 51-8011.00
51-8012.00 51-8092.00 51-8093.00 51-9022.00 51-9031.00 51-9071.00 51-9082.00 51-9083.00
51-9141.00 51-9191.00 51-9192.00 51-9193.00 51-9194.00 51-9195.03 51-9195.04 51-9195.05
51-9196.00 51-9198.00 49-9098.00
53-1041.00 53-1044.00 53-2012.00 53-2021.00 53-2022.00 53-3011.00 53-3053.00 53-4011.00
53-4013.00 53-4022.00 53-4031.00 53-5011.00 53-5021.00 53-5031.00 53-6032.00 53-6041.00
53-6051.01 53-6061.00 53-7031.00 53-7041.00 53-7071.00 53-7072.00 53-7121.00
""".split()

T4 = """
11-9179.01 13-1041.06 15-1299.03 17-2121.00 19-1031.02 21-1091.00
25-3031.00 25-9021.00 29-1071.00 29-1071.01 29-1125.00
31-9094.00 33-3052.00 35-2013.00 39-3012.00 39-3021.00 39-5093.00 41-2012.00 33-9031.00
43-2011.00 43-2021.00 43-9022.00 45-2091.00
47-5011.00 47-5012.00 47-5013.00 47-5041.00 47-5043.00 47-5044.00 47-5071.00 47-5081.00
49-9092.00 51-2061.00 51-9071.06 51-9151.00 53-5022.00 53-6011.00 53-7073.00
""".split()

data = json.load(open(DATA))
occ = data["occupations"]
tier_of = {}
for soc in T1:
    tier_of[soc] = 1
for soc in T3:
    tier_of[soc] = 3
for soc in T4:
    tier_of[soc] = 4

json.dump(tier_of, open(os.path.join(BASE, "build", "hu_tiers.json"), "w"))
unknown = [s for s in tier_of if s not in {c["soc"] for c in occ}]
for c in occ:
    c["huTier"] = tier_of.get(c["soc"], 2)
    c["include"] = c["huTier"] <= 2

data["meta"]["tiering"] = {
    "method": "szakértői becslés (nincs publikus FEOR-4 létszámadat)",
    "date": "2026-07-30",
    "counts": {t: sum(1 for c in occ if c["huTier"] == t) for t in (1, 2, 3, 4)},
}
json.dump(data, open(DATA, "w"), ensure_ascii=False, indent=1)

print("tier counts:", data["meta"]["tiering"]["counts"])
print("termékbe (T1+T2):", sum(1 for c in occ if c["include"]))
if unknown:
    print("FIGYELEM, nem létező SOC a listákban:", unknown)

# ── Excel review-export ─────────────────────────────────────────────────────
DIM_ORDER = ["INTE", "RESO", "TEMP", "ADAP", "THOR", "OPEN"]
DIM_SHORT = {"INTE": "H", "RESO": "E", "TEMP": "X", "ADAP": "A", "THOR": "C", "OPEN": "O"}
ENTRY_HU = {"open": "nincs formális követelmény", "course": "tanfolyam/rövid képzés",
            "vocational": "szakma/technikus", "higher": "felsőfokú diploma",
            "specialized": "szakirányú diploma + szakvizsga"}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Katalógus review"
headers = [
    "DÖNTÉS (marad/törlés)", "VÉGLEGES HU NÉV (ha módosítod)", "MEGJEGYZÉS",
    "tier", "HU név (auto)", "EN cím", "FEOR-08", "FEOR név", "ISCO-08", "ISCO név",
    "HU leírás", "Belépés", "Tipikus végzettség", "Holland top3",
    "R", "I", "A", "S", "E", "C",
    "HEXACO cél-profil (top4)", "H abs", "E abs", "X abs", "A abs", "C abs", "O abs",
    "név-review?", "SOC", "ESCO kód", "piaci nevek",
]
ws.append(headers)
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.fill = PatternFill("solid", start_color="EDE7DD")
ws.freeze_panes = "D2"

TIER_FILL = {1: "D8EFD8", 2: "FFFFFF", 3: "FFF3D6", 4: "F8D8D8"}

for c in sorted(occ, key=lambda x: (x["huTier"], x["isco08"] or "", x["huLabel"])):
    demand = " · ".join(
        f"{DIM_SHORT[e['dim']]} {e['target']:.0f}±{e['tolerance']:.0f} (w={e['weight']:.2f})"
        for e in c["demand"] if e["weight"] >= 0.10
    )
    ri = c["riasec"] or {}
    row = [
        "marad" if c["include"] else "törlés",
        "", "",
        c["huTier"], c["huLabel"], c["enTitle"],
        "/".join(c["feor08"]), c["feorLabel"], c["isco08"], c["iscoNameHu"],
        c["huDescription"] or c["enDescription"],
        ENTRY_HU.get(c["entryLevel"] or "", ""), c["eduMinHu"] or "", c["riasecTop"] or "",
        ri.get("R"), ri.get("I"), ri.get("A"), ri.get("S"), ri.get("E"), ri.get("C"),
        demand,
        *[c["absolute"][d]["level"] for d in DIM_ORDER],
        "" if c["nameConfident"] else "IGEN",
        c["soc"], c["escoCode"], ", ".join(c["huMarketLabels"][:5]),
    ]
    ws.append(row)
    fill = TIER_FILL[c["huTier"]]
    if fill != "FFFFFF":
        ws.cell(row=ws.max_row, column=4).fill = PatternFill("solid", start_color=fill)

widths = {"A": 18, "B": 26, "C": 22, "D": 6, "E": 34, "F": 40, "G": 10, "H": 30, "I": 9,
          "J": 30, "K": 60, "L": 22, "M": 28, "N": 11, "U": 52, "AB": 12, "AC": 12,
          "AD": 12, "AE": 26}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=11).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=21).alignment = Alignment(wrap_text=True, vertical="top")

# Módszertani lap
ws2 = wb.create_sheet("Módszer")
for line in [
    ["Foglalkozás-katalógus v2 — review tábla", ""],
    ["Generálva", "2026-07-30"],
    ["Tételszám", len(occ)],
    ["", ""],
    ["Tier", "Jelentés"],
    ["T1", "nagy létszámú, gyakori magyar foglalkozás"],
    ["T2", "létezik Magyarországon, közepes/kisebb létszám (default)"],
    ["T3", "niche: létezik, de nagyon szűk szegmens"],
    ["T4", "magyar munkaerőpiacon nem értelmezhető / elhanyagolható"],
    ["", ""],
    ["Mit tegyél", "Az A oszlopban írd át 'marad'/'törlés'-re; a B oszlopba a végleges magyar nevet, ha a javasolt nem jó."],
    ["", ""],
    ["Forrás — személyiség és érdeklődés", "O*NET 30.3 Database (U.S. Department of Labor/ETA), CC BY 4.0"],
    ["Forrás — magyar nevek, leírások", "ESCO v1.2 (Európai Bizottság)"],
    ["Forrás — FEOR-08 kódok", "KSH ISCO-08 → FEOR-08 fordítókulcs"],
    ["", ""],
    ["HEXACO cél-profil", "A 21 O*NET Work Style előjeles hatás-értékéből (Work Styles Impact, -3..+3) származtatva: foglalkozáson belüli centrálás + standardizálás + loading-mátrix. 'cél' = ideális érték 0-100, '±' = tolerancia, 'w' = súly."],
    ["HEXACO abs", "Ugyanaz centrálás nélkül: a szükséges ABSZOLÚT szint."],
    ["Holland", "O*NET Occupational Interests 1-7 skála 0-100-ra vetítve."],
    ["Belépés / végzettség", "O*NET Job Zone + a birtokosok által leggyakrabban jelölt végzettségi szint (AMERIKAI adat)."],
]:
    ws2.append(line)
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 110
for r in range(1, ws2.max_row + 1):
    ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws2.cell(row=1, column=1).font = Font(bold=True, size=13)

out = os.path.join(REPO, "docs/product/data/occupation-catalog-review.xlsx")
wb.save(out)
print("xlsx:", out, os.path.getsize(out), "bytes")
